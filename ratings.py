import copy
import time, os, re, openpyxl
import config
import mysql.connector
from mysql.connector import Error
from openpyxl.utils import get_column_letter


def difference(list1, list2):
    list_dif = [i for i in list1 + list2 if i not in list1 or i not in list2]
    return list_dif


mydb = mysql.connector.connect(
    host=config.DB_HOST,
    user=config.DB_USER_NAME,
    password=config.DB_PASSWORD,
    database=config.DB_NAME
)
mycursor = mydb.cursor()

# Iterate directory
for path in os.listdir(config.input_dir_path):
    try:
        if os.path.isfile(os.path.join(config.input_dir_path, path)):
            full_path = os.path.join(config.input_dir_path, path)
            file_name = os.path.splitext(os.path.basename(full_path))[0]
            print(file_name)
            # exit()

            current_time = int(time.time())
            output_file_path = config.output_dir_path + '\\' + file_name + "_result.xlsx"

            # Errors handling
            wb_error = openpyxl.Workbook()
            sheet_error = wb_error.active
            a1 = sheet_error['A1']
            a1.value = "Error"

            if not full_path.endswith('.xlsx'):
                print("Error occurs while reading input file")
                a1 = sheet_error['A2']
                a1.value = "Error occurs while reading input file"
                wb_error.save(output_file_path)
                print("File created successfully.")
                time.sleep(3)
                continue

            # workbook object is created
            wb_obj = openpyxl.load_workbook(full_path)

            # Write output to Excel file
            wb = openpyxl.Workbook()
            sheet = wb.active
            a1 = sheet['A1']
            a1.value = "Content ID"
            b1 = sheet['B1']
            b1.value = "Territory"
            c1 = sheet['C1']
            c1.value = "Subclass"
            d1 = sheet['D1']
            d1.value = "Expected Rating"
            e1 = sheet['E1']
            e1.value = "Derived Rating"
            f1 = sheet['F1']
            f1.value = "Expected Advisory"
            g1 = sheet['G1']
            g1.value = "Derived Advisory"
            h1 = sheet['H1']
            h1.value = "Derived Advisory Name"
            i1 = sheet['I1']
            i1.value = "Rating Status"
            j1 = sheet['J1']
            j1.value = "Advisory Status"

            for sheet_obj in wb_obj.worksheets:
                # print(sheet_obj.title)
                if sheet_obj.title in ['Ratings & Advisory', 'Advisory Calculation']:
                    break

                total_rows_count = len([row for row in sheet_obj if not all([cell.value is None for cell in row])])
                if total_rows_count > 0 or sheet_obj.title in ['Profanity', 'Questionnaire', 'Subclass Output']:
                    max_columns = sheet_obj.max_column
                    sheet_columns_dict = dict()
                    column = 1
                    while column <= max_columns:
                        column_value = str(
                            sheet_obj.cell(row=1, column=column).value).strip()
                        sheet_columns_dict[column_value] = column
                        column = column + 1

                    # print(sheet_columns_dict)
                    # exit()
                    # Title information
                    if sheet_obj.title == 'Title Info':
                        # continue  # Test
                        # print(sheet_columns_dict)
                        # exit()
                        row = 2

                        # Title information
                        title_type = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Title type*']).value)
                        if title_type == 'Episode':
                            title_type = 'Series'
                        feature_name = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Feature Name*']).value)
                        # Added timestamp to avoid content_already exists error
                        feature_name = feature_name + '_' + str(current_time)

                        series_name = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Series Name*']).value)
                        series_name = series_name + '_' + str(current_time)

                        series_number = sheet_obj.cell(row=row, column=sheet_columns_dict['Season Number*']).value
                        episode_number = sheet_obj.cell(row=row, column=sheet_columns_dict['Episode Number*']).value

                        episode_name = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Episode Name']).value)
                        if episode_name == 'None':
                            episode_name = ''

                        version = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Version*']).value)

                        ext_content_id = str(
                            sheet_obj.cell(row=row, column=sheet_columns_dict['External Content ID']).value)
                        if ext_content_id == 'None':
                            ext_content_id = ''  # For now
                            ext_content_id = str(current_time)
                        else:
                            ext_content_id = ext_content_id + "_" + str(current_time)

                        imdb_id = str(sheet_obj.cell(row=row, column=sheet_columns_dict['IMDB ID']).value)
                        if imdb_id == 'None':
                            imdb_id = ''

                        country_of_origin = str(
                            sheet_obj.cell(row=row, column=sheet_columns_dict['Country of Origin*']).value)
                        original_language = str(
                            sheet_obj.cell(row=row, column=sheet_columns_dict['Original Language*']).value)

                        runtime = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Runtime*']).value)
                        release_year = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Release Year*']).value)
                        synopsis = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Synopsis']).value)

                        writer = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Writer*']).value)
                        director = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Director*']).value)
                        producer = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Producer*']).value)
                        productionCompany = str(
                            sheet_obj.cell(row=row, column=sheet_columns_dict['Production Company*']).value)
                        cast = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Cast*']).value)

                        us_rating = str(sheet_obj.cell(row=row, column=sheet_columns_dict['US Rating']).value)
                        if us_rating == 'None':
                            us_rating = ''

                        us_rating_advisory = str(
                            sheet_obj.cell(row=row, column=sheet_columns_dict['US Rating Advisory']).value)
                        if us_rating_advisory == 'None':
                            us_rating_advisory = ''

                        artwork = str(sheet_obj.cell(row=row, column=sheet_columns_dict['Artwork']).value)
                        if artwork == 'None':
                            artwork = ''

                        original_air_date = sheet_obj.cell(row=row,
                                                           column=sheet_columns_dict['Original Air Date']).value
                        if original_air_date is None:
                            original_air_date = ''
                        else:
                            original_air_date = original_air_date.strftime("%m/%d/%Y")

                        print(original_air_date)
                        # exit()
                        # Territories
                        territories = str(
                            sheet_obj.cell(row=row, column=sheet_columns_dict['Territories*']).value).split(',')

                        advisory_business_rule = str(
                            sheet_obj.cell(row=row, column=sheet_columns_dict['Advisory Business Rule']).value)

                        # print(territories)
                        # exit()

                        if title_type == 'Feature':
                            print('Feature')
                            args = [0, 0, title_type, feature_name, title_type, title_type, feature_name, 0, 0, '',
                                    '', '', version, original_language, country_of_origin, release_year, runtime,
                                    writer, director, producer, cast, productionCompany, us_rating,
                                    us_rating_advisory, ext_content_id, '', '', imdb_id, artwork, original_air_date,
                                    synopsis, 'user2', 23, 0]
                        else:
                            print('Series')
                            args = [0, 0, title_type, series_name, 'Episode', title_type, series_name, series_number,
                                    episode_number, '',
                                    '', '', version, original_language, country_of_origin, release_year, runtime,
                                    writer, director, producer, cast, productionCompany, us_rating,
                                    us_rating_advisory, ext_content_id, '', '', imdb_id, artwork, original_air_date,
                                    synopsis, 'user2', 23, 0]
                        try:
                            print(args)
                            result_args = mycursor.callproc('SP_SaveTitleAndContent', args)
                            mydb.commit()
                            print(result_args)
                            # exit()

                            title_sp_response = result_args[33].split(',')
                            print(title_sp_response)
                            # exit()
                            result = ''
                            # Check if data saved successfully
                            if int(title_sp_response[0]) in [10, 12]:
                                for result in mycursor.stored_results():
                                    result = result.fetchall()

                                content_id = result[0][0]
                                title_id = result[0][1]
                                print("content_id :" + str(content_id))
                                print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^ title_id :" + str(title_id))
                                print(type(title_id))
                            else:
                                a1 = sheet_error['A2']
                                a1.value = title_sp_response[1]
                                wb_error.save(output_file_path)
                                print("File created successfully.")
                                time.sleep(3)
                                break
                        except Error as e:
                            print(e)
                            a1 = sheet_error['A2']
                            a1.value = str(e)
                            wb_error.save(output_file_path)
                            print("File created successfully.")
                            time.sleep(3)
                            break

                    elif sheet_obj.title == 'Questionnaire':
                        # continue  # Testing
                        # content_id =1589
                        # title_id = 1565
                        # title_type='Feature'
                        # print(sheet_columns_dict['Selected Answer'])
                        # exit()
                        row = 2
                        answers_dict = dict()
                        while row <= total_rows_count:
                            selected_answer = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Selected Answers']).value)
                            if selected_answer == 'Yes':
                                question_id = str(
                                    sheet_obj.cell(row=row, column=sheet_columns_dict['Question Id']).value)
                                option_id = str(
                                    sheet_obj.cell(row=row, column=sheet_columns_dict['Option Id']).value)
                                print("question_id : " + str(question_id))
                                print("option_id :" + str(option_id))
                                if question_id not in answers_dict.keys():
                                    answers_dict[question_id] = []
                                answers_dict[question_id].append(option_id)
                            row += 1
                        print(answers_dict)
                        # exit()

                    elif sheet_obj.title == 'Profanity':
                        # continue  # Testing
                        # print(sheet_obj.title)
                        # exit()
                        # content_id = 2
                        # title_id = 1
                        # answers_dict = {}  # Test
                        row = 2
                        profanity_dict = dict()
                        while row <= total_rows_count:
                            selected_answer = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Selected Answer']).value)
                            if selected_answer == 'Yes':
                                question_id = str(
                                    sheet_obj.cell(row=row, column=sheet_columns_dict['question_id_word']).value)
                                option1 = str(
                                    sheet_obj.cell(row=row, column=sheet_columns_dict['sub_option_id_1']).value)
                                option2 = str(
                                    sheet_obj.cell(row=row, column=sheet_columns_dict['sub_option_id_2']).value)
                                option3 = str(
                                    sheet_obj.cell(row=row, column=sheet_columns_dict['sub_option_id_3']).value)

                                if option1 == '0':
                                    option1 = None
                                if option2 == '0':
                                    option2 = None
                                if option3 == '0':
                                    option3 = None

                                print("question_id : " + str(question_id))
                                print("option1 :" + str(option1))
                                print("option2 :" + str(option2))
                                print("option3 :" + str(option3))

                                if question_id not in profanity_dict.keys():
                                    profanity_dict[question_id] = []
                                profanity_dict[question_id].append(option1)
                                profanity_dict[question_id].append(option2)
                                profanity_dict[question_id].append(option3)
                            row += 1
                        print(profanity_dict)
                        # exit()

                        # Loop the dict and call SP_SaveAnswers SP to save the answers
                        if len(answers_dict) > 0:
                            for question in answers_dict.keys():
                                answer = ','.join(answers_dict[question])
                                args = [title_id, content_id, question, answer, 'user2', 23, 0]
                                print('^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^', args)
                                print(type(title_id))
                                result_args = mycursor.callproc('SP_SaveAnswers', args)
                                mydb.commit()
                                print(result_args)
                                # exit()
                        if len(profanity_dict) > 0:
                            for option_id_word in profanity_dict.keys():
                                sub_option_1 = profanity_dict[option_id_word][0]
                                sub_option_2 = profanity_dict[option_id_word][1]
                                sub_option_3 = profanity_dict[option_id_word][2]
                                args = [title_id, content_id, option_id_word, sub_option_1, sub_option_2, sub_option_3,
                                        'user2', 23, 0]
                                print('^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^', args)
                                print(type(title_id))
                                result_args = mycursor.callproc('SP_SaveAnswersProfanity', args)
                                mydb.commit()
                                print(result_args)
                                # exit()
                        if len(answers_dict) <= 0 and len(profanity_dict) <= 0:
                            print("No question answered so exit the loop")
                            a1 = sheet_error['A2']
                            a1.value = "No questions are answered"
                            wb_error.save(output_file_path)
                            print("File created successfully.")
                            time.sleep(3)
                            break

                        time.sleep(5)
                        # Call rating request SP
                        print("CAll SP_RaiseRatingsRequest")
                        # territories = ['India']  # Test
                        # str_territory = ', '.join(f'"{territory}"' for territory in territories)
                        str_territory = ', '.join(f'"{territory.strip().lower()}"' for territory in territories)
                        # Get territoryID
                        sql_territory = "SELECT id FROM m_territory where territory IN (" + str_territory + ")"
                        print(sql_territory)
                        # exit()
                        mycursor.execute(sql_territory)
                        sql_territory_result = mycursor.fetchall()
                        territory_id_arr = []
                        for territory_id in sql_territory_result:
                            territory_id_arr.append(next(iter(territory_id)))

                        print(territory_id_arr)
                        str_territory_id = ', '.join(f'{territory}' for territory in territory_id_arr)
                        print(str_territory_id)
                        args = [content_id, title_type, str_territory_id, 1, 1, 3, 'user2', 23, 0]  # 1 is pre-order, 2 is ordered, 3 is submitted, 4 is delivered
                        # args = [content_id, title_type, str_territory_id, 1, 1, 'Submitted', 'user2', 23, 0]
                        # args = [content_id, 'Feature', str_territory_id, 1, 1, 'Submitted', 'user2', 23, 0]
                        print(args)

                        result_args = mycursor.callproc('SP_RaiseRatingsRequest', args)
                        mydb.commit()
                        print(result_args)
                        time.sleep(3)
                        # exit()
                        if result_args[8] == 'Ratings requested successfully':
                            # Call CalculateAndSaveRatings to generate rating
                            args = [content_id, 0]
                            print('^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^ ------------------------------', args)
                            print(type(args[0]))
                            result_args = mycursor.callproc('SP_CalculateAndSaveRatings', args)
                            # result_args = mycursor.callproc('SP_CalculateAndSaveRatings', args)
                            # result_args = mycursor.callproc('zratings', args)
                            print('##########################  SP_CalculateAndSaveRatings')
                            print(result_args)
                            mydb.commit()
                            time.sleep(3)
                            # exit()
                            if result_args[1] != 'success':
                                a1 = sheet_error['A2']
                                a1.value = result_args[1]
                                wb_error.save(output_file_path)
                                print("File created successfully.")
                                time.sleep(3)
                                break
                        else:
                            a1 = sheet_error['A2']
                            a1.value = result_args[8]
                            wb_error.save(output_file_path)
                            print("File created successfully.")
                            time.sleep(3)
                            break

                    elif sheet_obj.title == 'Title Output':
                        # print(output_columns_dict)
                        # print(output_columns_dict['Territory'])

                        # content_id = 2726  # Test
                        # title_id = 2383  # Test
                        row = 2
                        args = [content_id, 'title', 23, 0]
                        result_args = mycursor.callproc('SP_GetRatingsForContent', args)
                        print(result_args)
                        # exit()
                        if result_args[3] == 'success':
                            record_counter = 1
                            for result in mycursor.stored_results():
                                if record_counter == 3:
                                    db_rating_results = result.fetchall()
                                    break
                                record_counter += 1
                        else:
                            a1 = sheet_error['A2']
                            a1.value = result_args[3] + "(" + str(content_id) + ")"
                            wb_error.save(output_file_path)
                            print("File created successfully.")
                            time.sleep(3)
                            break

                        print(db_rating_results)
                        # exit()

                        while row <= total_rows_count:
                            territory = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Territory']).value).strip()
                            rating = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Rating Code']).value).strip()
                            advisory = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Advisory']).value).strip()
                            if advisory == 'None':
                                advisory = ''

                            if rating == 'None':
                                rating = ''

                            if territory == 'None':
                                territory = ''

                            excel_advisory_code_arr = []
                            if advisory != 'None':
                                excel_advisory_code = advisory.split(",")
                                for adv_code in excel_advisory_code:
                                    if "-" in adv_code:
                                        excel_advisory_code_arr.append(adv_code.split("-")[0].strip())

                            print("Excel territory :" + territory)
                            print("Excel rating :" + str(rating))
                            print(excel_advisory_code_arr)

                            rating_status = advisory_status = 'Fail'
                            db_territory = db_rating = db_rating_territory = ''
                            db_advisory_code_territory = db_advisory_name_territory = ''
                            db_advisory_status_territory = ''

                            for db_rating_result in db_rating_results:
                                db_territory = db_rating_result[8]
                                db_rating = db_rating_result[14]
                                db_advisory_code = db_rating_result[19]
                                db_advisory_name = db_rating_result[20]

                                db_advisory_code_arr = []

                                if db_advisory_code is None:
                                    db_advisory_code = ''

                                if db_advisory_name is None:
                                    db_advisory_name = ''

                                if db_advisory_code != '':
                                    db_advisory_code_arr = db_advisory_code.replace(' ', '').split(';')

                                print("DB territory :" + db_territory)
                                print("DB rating :" + str(db_rating))

                                print(db_advisory_code_arr)
                                print(db_advisory_name)
                                # exit()
                                z = difference(db_advisory_code_arr, excel_advisory_code_arr)

                                territory_matched = False
                                if str(territory).lower() == str(db_territory).lower():
                                    territory_matched = True
                                    if str(db_rating).lower() == str(rating).lower():
                                        rating_status = 'Pass'
                                        print(db_territory + "<<" + rating_status)
                                    if not z and (len(db_advisory_code_arr) == len(excel_advisory_code_arr)):
                                        advisory_status = 'Pass'
                                        print(db_territory + ">>" + advisory_status)
                                    break
                                else:
                                    continue  # Go to next db record

                            if territory_matched == False:
                                db_advisory_code = db_advisory_name = db_rating = ''
                                db_advisory_code_arr = []
                                if str(db_rating).lower() == str(rating).lower():
                                    rating_status = 'Pass'
                                    advisory_status = 'Pass'

                                if not z and (len(db_advisory_code_arr) == len(excel_advisory_code_arr)):
                                    advisory_status = 'Pass'

                            a1 = sheet['A' + str(row)]
                            a1.value = content_id

                            b1 = sheet['B' + str(row)]
                            b1.value = territory.capitalize()

                            c1 = sheet['C' + str(row)]
                            c1.value = 'Title'

                            d1 = sheet['D' + str(row)]
                            d1.value = rating

                            e1 = sheet['E' + str(row)]
                            e1.value = db_rating
                            # e1.value = ''

                            f1 = sheet['F' + str(row)]
                            f1.value = advisory

                            g1 = sheet['G' + str(row)]
                            g1.value = db_advisory_code
                            # g1.value = ''

                            h1 = sheet['H' + str(row)]
                            h1.value = db_advisory_name
                            # g1.value = ''

                            i1 = sheet['I' + str(row)]
                            i1.value = rating_status

                            j1 = sheet['J' + str(row)]
                            j1.value = advisory_status

                            row += 1
                            cont_row = copy.deepcopy(row)

                        # wb.save(output_file_path)
                        # print("File created successfully.")
                        # time.sleep(3)
                        # # exit()

                    elif sheet_obj.title == 'Subclass Output':
                        row = 2
                        args = [content_id, 'subclass', 23, 0]
                        result_args = mycursor.callproc('SP_GetRatingsForContent', args)
                        print(result_args)
                        # exit()
                        if result_args[3] == 'success':
                            record_counter = 1
                            for result in mycursor.stored_results():
                                if record_counter == 3:
                                    db_rating_results = result.fetchall()
                                    break
                                record_counter += 1
                        else:
                            a1 = sheet_error['A2']
                            a1.value = result_args[3] + "(" + str(content_id) + ")"
                            wb_error.save(output_file_path)
                            print("File created successfully.")
                            time.sleep(3)
                            break

                        print(db_rating_results)
                        # exit()

                        while row <= total_rows_count:
                            territory = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Territory']).value).strip()
                            subclass = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Subclass']).value).strip()
                            rating = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Subclass rating']).value).strip()
                            advisory = str(
                                sheet_obj.cell(row=row, column=sheet_columns_dict['Subclass advisory']).value).strip()
                            if advisory == 'None':
                                advisory = ''

                            if subclass == 'None':
                                subclass = ''

                            if rating == 'None':
                                rating = ''

                            if territory == 'None':
                                territory = ''

                            excel_advisory_code_arr = []
                            if advisory != 'None':
                                excel_advisory_code = advisory.split(",")
                                for adv_code in excel_advisory_code:
                                    # if "-" in adv_code:
                                    excel_advisory_code_arr.append(adv_code.split("-")[0].strip())

                            print("Excel territory :" + territory)
                            print("Excel subclass :" + subclass)
                            print("Excel rating :" + str(rating))
                            excel_advisory_code_arr = list(filter(None, excel_advisory_code_arr))
                            print(excel_advisory_code_arr)

                            rating_status = advisory_status = 'Fail'
                            db_territory = db_rating = db_rating_territory = ''
                            db_advisory_code_territory = db_advisory_name_territory = ''
                            db_advisory_status_territory = ''

                            for db_rating_result in db_rating_results:
                                db_territory = db_rating_result[8]
                                db_subcategory = db_rating_result[13]
                                db_rating = db_rating_result[14]
                                db_advisory_code = db_rating_result[19]
                                db_advisory_name = db_rating_result[20]

                                db_advisory_code_arr = []

                                if db_advisory_code is None:
                                    db_advisory_code = ''

                                if db_advisory_name is None:
                                    db_advisory_name = ''

                                if db_advisory_code != '':
                                    db_advisory_code_arr = db_advisory_code.replace(' ', '').split(';')

                                print("DB territory :" + db_territory)
                                print("DB rating :" + str(db_rating))

                                print(db_advisory_code_arr)
                                print(db_advisory_name)
                                # exit()
                                z = difference(db_advisory_code_arr, excel_advisory_code_arr)

                                territory_matched = False
                                if str(territory).lower() == str(db_territory).lower() and str(subclass).lower() == str(db_subcategory).lower():
                                    territory_matched = True
                                    if str(db_rating).lower() == str(rating).lower():
                                        rating_status = 'Pass'
                                        print(db_territory + "<<" + rating_status)
                                    if not z and (len(db_advisory_code_arr) == len(excel_advisory_code_arr)):
                                        advisory_status = 'Pass'
                                        print(db_territory + ">>" + advisory_status)
                                    break
                                else:
                                    continue  # Go to next db record

                            if territory_matched == False:
                                db_advisory_code = db_advisory_name = db_rating = ''
                                db_advisory_code_arr = []
                                if str(db_rating).lower() == str(rating).lower():
                                    rating_status = 'Pass'
                                    advisory_status = 'Pass'

                                if not z and (len(db_advisory_code_arr) == len(excel_advisory_code_arr)):
                                    advisory_status = 'Pass'

                            a1 = sheet['A' + str(cont_row)]
                            a1.value = content_id

                            b1 = sheet['B' + str(cont_row)]
                            b1.value = territory.capitalize()

                            c1 = sheet['C' + str(cont_row)]
                            c1.value = subclass

                            d1 = sheet['D' + str(cont_row)]
                            d1.value = rating

                            e1 = sheet['E' + str(cont_row)]
                            e1.value = db_rating
                            # e1.value = ''

                            f1 = sheet['F' + str(cont_row)]
                            f1.value = advisory

                            g1 = sheet['G' + str(cont_row)]
                            g1.value = db_advisory_code
                            # g1.value = ''

                            h1 = sheet['H' + str(cont_row)]
                            h1.value = db_advisory_name
                            # g1.value = ''

                            i1 = sheet['I' + str(cont_row)]
                            i1.value = rating_status

                            j1 = sheet['J' + str(cont_row)]
                            j1.value = advisory_status

                            cont_row += 1
                            row += 1

                        wb.save(output_file_path)
                        print("File created successfully.")
                        time.sleep(3)
                        # exit()
                else:
                    # Write output file with error message and break
                    # No data found for Title Info OR Questionniare
                    a1 = sheet_error['A2']
                    a1.value = "No data found for " + sheet_obj.title
                    wb_error.save(output_file_path)
                    print("File created successfully.")
                    time.sleep(3)
                    break

    except Error as e:
        print(e)
        print("Error occurs while reading input file")
        a1 = sheet_error['A2']
        a1.value = "Error occurs while reading input file"
        wb_error.save(output_file_path)
        print("File created successfully.")
        time.sleep(3)
        continue

mycursor.close()
mydb.close()
